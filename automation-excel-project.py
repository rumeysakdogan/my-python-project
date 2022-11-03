import openpyxl
"""
Learn how to work with Spreadsheets
- Read spreadsheet file and automate stuff
- we have inventory.xlsx file
Ex1: List each company with respective product count
Ex2: List products with inventory less than 10
Ex3: List each company with respective total inventory value
Ex4: Write to spreadsheet: Calculate and write inventory value for each product into spreadsheet
"""

inv_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inv_file["Sheet1"]

products_per_supplier = {}  # { "company_name" : "num_of_products", ...}
total_val_per_supplier = {}
products_under_10_inv = {}

for product_row in range(2,product_list.max_row + 1):  # [2, 3, 4, 75]
    supplier_name = product_list.cell(product_row, 4).value
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    product_num = product_list.cell(product_row, 1).value
    inventory_price = product_list.cell(product_row, 5)

    # calculation number of products per supplier
    if supplier_name in products_per_supplier:
        current_num_products = products_per_supplier.get(supplier_name)
        products_per_supplier[supplier_name] = current_num_products + 1
    else:
        print("adding a new supplier")
        products_per_supplier[supplier_name] = 1

    # calculation total value of inventory per supplier
    if supplier_name in total_val_per_supplier:
        current_total_value = total_val_per_supplier.get(supplier_name)
        total_val_per_supplier[supplier_name] = current_total_value + inventory * price
    else:
        total_val_per_supplier[supplier_name] = inventory * price

    # logic for product with inventory less than 10
    if inventory < 10:
        products_under_10_inv[int(product_num)] = int(inventory)

    # add value for total inventory price
    inventory_price.value = inventory * price

print(products_per_supplier)
print(total_val_per_supplier)
print(products_under_10_inv)

inv_file.save("inventory_with_total_value.xlsx")


"""
Different ways to work with files
-> Python has several built-in functions for handling files in general
ex: io module ->> create, read, write
-> Python package to work with spreadsheets specifically
 * easier to use and more practical functions for spreadsheets specifically
"""

# range():creates a sequence of numbers
# starting from 0 by default
# ex: range(75) --> [0, 1, 2, 3 .., 74]




